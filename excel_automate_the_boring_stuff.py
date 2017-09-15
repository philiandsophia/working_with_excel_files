# -*- coding: utf-8 -*-
"""
Created on Thu Sep  7 16:38:44 2017

@author: choip
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font


wb = openpyxl.load_workbook('example.xlsx')
#print (wb.get_sheet_names())
#print (wb.get_sheet_by_name("Sheet3"))
#anotherSheet = wb.active
#print (anotherSheet)
#print (anotherSheet.title)

sheet = wb.get_sheet_by_name('Sheet1')
#print (sheet['A1'].value)
#c = sheet['B1']
#print (c.value)
#print ('Row ' + str(c.row) + ', Column '+c.column + ' is ' + c.value)
#print ('Cell ' + c.coordinate + ' is ' + c.value)
#
#print (sheet['C1'].value)

#for i in range(1,8,2):
    #print (i, sheet.cell(row = i, column=2).value)
#first row or column is rep. by integer 1, not 0

#print (sheet.max_row)
#print (sheet.max_column)

#print (get_column_letter(1))
#print (get_column_letter(2))
#print (get_column_letter(4774))
#print (get_column_letter(sheet.max_column))
#
#print (column_index_from_string('A'))
#print (column_index_from_string('GAH'))
#in excel spreadsheet, columns are represented by alphabets
#for example, first column is A, second column is B and 27th is AA

#print (tuple(sheet['A1':'C3']))
#get all celss in row A, B and C from column 1 to column 3

#for rowOfCellobjects in sheet['A1':'C3']:
#    for cellObj in rowOfCellobjects:
#        print (cellObj.coordinate,cellObj.value)
#    print ('--- END OF ROW ---')

alist = []

#for cellObj in sheet.columns:
    #alist.append(cellObj)
    #for e in alist[1]:
        #print (e.value, type(e.value))
#sheet columns is a generator that retruns all the values of all the columns
#in tuples(each column gets its own tuple)
#that's why I had to use a convoulted mehtod    


#for e in alist[1]:
        #print (e.value, type(e.value))


wb = openpyxl.Workbook()
print (wb.get_sheet_names)
sheet = wb.active
print (sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print (wb.get_sheet_names())
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')
#you need to save the worksheet

#always save the new edited spreadsheet to a different filename than original
#as codes can mess up the data

wb = openpyxl.Workbook()

wb.create_sheet('First Sheet',0)

print (wb.get_sheet_names())
wb.create_sheet('Middle Sheet',1)
print (wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
#removes sheet
print (wb.get_sheet_names())
sheet = wb.get_sheet_by_name('First Sheet')
sheet['A1']= 'Hello word!'
print (sheet['A1'].value)

italic24Font = Font(size = 24, italic = True)
sheet['A1'].font = italic24Font
wb.save('styled.xlsx')

#Table 12-2. Keyword Arguments for Font style Attributes
#
#Keyword argument
#
#Data type
#
#Description
#
#name
#
#String
#
#The font name, such as 'Calibri' or 'Times New Roman'
#
#size
#
#Integer
#
#The point size
#
#bold
#
#Boolean
#
#True, for bold font
#
#italic
#
#Boolean
#
#True, for italic font

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

fontObj1 = Font(name = 'Times New Roman' ,bold = True)

sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size = 24, italic = True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

wb.save('styles.xlsx')

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=Sum(A1:A2)'
wb.save('writeFormula.xlsx')

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelev cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells'
wb.save('merged.xlsx')

wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')
wb.save('merged.xlsx')


from openpyxl import Workbook
wb = Workbook()
ws = wb.active
for i in range(10):
    ws.append([i])

from openpyxl.chart import BarChart, Reference, Series
values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values) 
ws.add_chart(chart, "E15")
wb.save("SampleChart.xlsx")

